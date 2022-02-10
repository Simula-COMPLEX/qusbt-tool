import os
from openpyxl import Workbook
from gtime import GlobalTime
from testing import dec2bin

import numpy as np
from jmetal.algorithm.multiobjective import RandomSearch
from jmetal.algorithm.singleobjective import GeneticAlgorithm
from jmetal.operator import IntegerPolynomialMutation
from jmetal.operator import IntegerSBXCrossover
from jmetal.operator.selection import BestSolutionSelection
from jmetal.util.termination_criterion import StoppingByEvaluations

from TestingProblem import TestingProblem

from GenerateUnitTest import GenerateUnitTest

import sys
if sys.platform.startswith('win32'):
    ROOT='\\'
elif sys.platform.startswith('linux'):
    ROOT='/'
elif sys.platform.startswith('darwin'):
    ROOT='/'

class Search:


    @staticmethod
    def search(config_dic):
        wb = Workbook()
        solution_sheet = wb.active
        solution_sheet.title = 'Best Solution'
        log_sheet = wb.create_sheet('Log Information')
        solution_sheet['A4'] = 'Simulation Time'
        solution_sheet['B4'].value = 0

        problem = TestingProblem(config_dic, solution_sheet, log_sheet)
        np.set_printoptions(linewidth=500)

        if 'mutation_probability' in config_dic.keys():
            algorithm = GeneticAlgorithm(
                problem=problem,
                population_size=config_dic['population_size'],  # 10
                offspring_population_size=config_dic['offspring_population_size'], #10
                mutation=IntegerPolynomialMutation(probability=config_dic['mutation_probability'], distribution_index=config_dic['mutation_distribution_index']),
                crossover=IntegerSBXCrossover(probability=config_dic['crossover_probability'], distribution_index=config_dic['crossover_distribution_index']),  # 0.9, 20
                selection=BestSolutionSelection(),
                termination_criterion=StoppingByEvaluations(max_evaluations=config_dic['max_evaluations'])  # 500
            )
        else:
            algorithm = GeneticAlgorithm(
                problem=problem,
                population_size=config_dic['population_size'],  # 10
                offspring_population_size=config_dic['offspring_population_size'], #10
                mutation=IntegerPolynomialMutation(probability=1.0/problem.number_of_variables, distribution_index=config_dic['mutation_distribution_index']),
                crossover=IntegerSBXCrossover(probability=config_dic['crossover_probability'], distribution_index=config_dic['crossover_distribution_index']),  # 0.9, 20
                selection=BestSolutionSelection(),
                termination_criterion=StoppingByEvaluations(max_evaluations=config_dic['max_evaluations'])  # 500
            )

        algorithm.run()

        result = algorithm.get_result()
        result_int = result.variables
        solution = []
        for l in range(len(result_int)):
            solution.append(dec2bin(result_int[l],len(config_dic['inputID'])))

        #get count_times
        valid_input = config_dic['valid_input']
        pt = config_dic['p']
        count = []
        for l in range(len(solution)):
            count_times = 0
            for i in range(len(valid_input)):
                if valid_input[i] == solution[l] and pt[l] > 0:
                    count_times += 1
            count.append(count_times*100)

        GenerateUnitTest.generateUnitTestClass(config_dic['module_name'], config_dic['inputID'], config_dic['outputID'], config_dic['num_qubit'], result_int, solution, count, config_dic['program_folder'],ROOT)


        solution.insert(0, 'Solution')
        obj = [-result.objectives[0][0]]

        obj.insert(0,'Fitness')
        total_time = [algorithm.total_computing_time]
        total_time.insert(0, 'Total time')

        for l in range(len(solution)):
            solution_sheet.cell(row=1,column=l+1,value=solution[l])
        for l in range(len(obj)):
            solution_sheet.cell(row=2,column=l+1,value=obj[l])
        for l in range(len(total_time)):
            solution_sheet.cell(row=3,column=l+1,value=total_time[l])

        wb.save(config_dic['program_folder']+ ROOT +config_dic['module_name']+'_result.xlsx')

        print("\n")
        print(f'Computing time: ${algorithm.total_computing_time}')
        print('Files '+config_dic['module_name']+'_result.xlsx'+' and '+config_dic['module_name']+'TEST.py'+' are generated.')

